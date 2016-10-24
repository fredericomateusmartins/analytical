#!/usr/bin/env python

from __future__ import print_function

import sys

from calendar import monthrange, weekday
from collections import OrderedDict
from datetime import date, datetime
from os import chdir, makedirs, path
from PIL import Image
from queue import Empty, Queue
from sys import argv, executable, exit
from time import sleep
from threading import Lock, Thread

from win32com.shell import shell # Windows Modules

from PyQt5 import QtCore, QtGui, QtWidgets # GUI Modules
from analyticalGUI import Ui_Interface

from googleapiclient import sample_tools # Analytics Modules
from googleapiclient.errors import HttpError
from oauth2client.client import AccessTokenRefreshError

from openpyxl import styles, workbook # Excel Modules
from openpyxl.chart import BarChart, LineChart, PieChart, series, Reference
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side
from openpyxl.worksheet import ColumnDimension 

from reportlab.graphics.charts.barcharts import VerticalBarChart # PDF Modules
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import LegendedPie
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase.pdfmetrics import registerFont, stringWidth
from reportlab.platypus.doctemplate import BaseDocTemplate, PageTemplate
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, PageBreak, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.rl_config import defaultPageSize


frozen = True if getattr(sys, 'frozen', False) else False # File state
curdir = argv[0] if frozen else __file__

realpath = argv[0].rpartition('\\')[0] # Important paths
folderpath = path.join(realpath, 'Reports')
imagespath = path.join(realpath, 'Images')
otherspath = path.join(realpath, 'Others')
infopath = path.join(realpath, 'companyinfo')
secretpath = path.join(realpath, 'client_secrets.json')

months = {'01':'Janeiro', '02':'Fevereiro', '03':'Março', '04':'Abril', '05':'Maio', '06':'Junho', '07':'Julho', '08':'Agosto', '09':'Setembro', '10':'Outubro', '11':'Novembro', '12':'Dezembro'}
weekdays = {0:'Segunda-Feira', 1:'Terça-Feira', 2:'Quarta-Feira', 3:'Quinta-Feira', 4:'Sexta-Feira', 5:'Sábado', 6:'Domingo'}

titles = ['Sessões (Gerais)', 'Fontes de Acesso', 'Palavras-Chaves', 'Sessões (Países)', 'Sessões (Cidades)', 'Páginas', 'Sessões (Diárias)', 'Tracking de Páginas', 'Sessões (Anuais)']

group = OrderedDict() # Google Analytics queries
group['session'] = ['ga:date', 'ga:sessions,ga:users,ga:pageviews,ga:uniquePageviews,ga:avgSessionDuration,ga:avgTimeOnPage,ga:bounceRate']
group['access'] = ['ga:source', 'ga:sessions', '-ga:sessions']
group['search'] = ['ga:keyword', 'ga:sessions', '-ga:sessions']
group['country'] = ['ga:country', 'ga:sessions, ga:pageviews, ga:bounceRate', '-ga:sessions']
group['city'] = ['ga:city', 'ga:sessions, ga:pageviews, ga:bounceRate', '-ga:sessions']
group['page'] = ['ga:pageTitle', 'ga:pageviews, ga:avgTimeOnPage', '-ga:pageviews']
group['daily'] = ['ga:dateHour', 'ga:users']
group['tracking'] = ['ga:landingPagePath, ga:secondPagePath', 'ga:entrances', '-ga:entrances']
group['yearly'] = ['ga:month', 'ga:sessions']

images = OrderedDict() # Images paths
images['logo'] = {'path':path.join(imagespath, 'logo.jpg'), 'width':180}
images['background'] = {'path':path.join(imagespath, 'background.bmp'), 'width':400}
images['header'] = {'path':path.join(imagespath, 'header.png'), 'width':460}
images['footer'] = {'path':path.join(imagespath, 'footer.png'), 'width':460}

registerFont(TTFont('Calibri', 'Calibri.ttf'))
registerFont(TTFont('Calibri-Bold', 'calibrib.ttf'))

headerfont = Font(name = 'Calibri', size = 12, bold = True, color = 'FF000000') # Excel fonts
align = Alignment(horizontal = 'center', vertical = 'center', text_rotation = 0, wrap_text = False, shrink_to_fit = False, indent = 0)
border = Border(left = Side(border_style = 'thick', color = 'FF000000'), right = Side(border_style = 'thick', color = 'FF000000'), top = Side(border_style = 'thick', color = 'FF000000'), bottom = Side(border_style = 'thick', color = 'FF000000'))

pH = ParagraphStyle(name = 'Header', fontName = 'Calibri-Bold', fontSize = 13, leftIndent = 20, firstLineIndent = -20, spaceBefore = 10, leading = 16) # PDF fonts
sH = ParagraphStyle(name = 'SubHeader', fontName = 'Calibri', fontSize = 12, leftIndent = 40, firstLineIndent = -20, spaceBefore = 5, leading = 16)
pT = ParagraphStyle(name = 'Title', fontName = 'Calibri-Bold', fontSize = 14, leftIndent = 20, firstLineIndent = -20, spaceBefore = 10, leading = 16) 
sT = ParagraphStyle(name = 'SubTitle', fontName = 'Calibri', fontSize = 13, leftIndent = 20, firstLineIndent = -20, spaceBefore = 10, leading = 16)
nS = ParagraphStyle(name = 'NormalSize', fontName = 'Calibri', fontSize = 11, leftIndent = 0, firstLineIndent = 0, spaceBefore = 7, leading = 16)
sS = ParagraphStyle(name = 'SmallSize', fontName = 'Calibri', fontSize = 7, leftIndent = 0, firstLineIndent = 0, spaceBefore = 5, leading = 16)


class Interface(QtWidgets.QMainWindow, Ui_Interface): # Main Interface

	def __init__(self):

		QtWidgets.QDialog.__init__(self)
		self.setupUi(self)
		self.retranslateUi(self)
		self.setDisabled(True)

		if not shell.IsUserAnAdmin() and frozen: # Informative box (Administration Access)
			self.warning = QtWidgets.QMessageBox()
			self.warning.information(self, 'Analytical', "\nThe program needs to be executed with administration privileges, please re-run the program.\n")
			raise SystemExit

		self.count = 0
		self.lengthbar = 0
		self.prevdone = False
		self.processgoing = False

		self.pipe = Queue()
		self.WriteTimer = QtCore.QTimer()
		self.SingleTimer = QtCore.QTimer()
		self.AnimationTimer = QtCore.QTimer()
		self.charlist = ['. .', '...', '.:.', ':::', ':.:', '...', ' . ', '   ']

		self.FetchProcess = Thread(target = OnLoadWorker, args = (self.pipe,))
		self.CurrentMonth = [date.today().strftime('%Y'), date.today().strftime('%m')]
		self.LastDay = monthrange(int(self.CurrentMonth[0]), int(self.CurrentMonth[1]))[1]

		self.SelectAll.setShortcut('Ctrl+S')
		self.SelectAll.setStatusTip('Ctrl+S')
		self.WriteButton.setShortcut('Ctrl+W')
		self.WriteButton.setStatusTip('Ctrl+W')
		self.SearchButton.setShortcut('Ctrl+F')
		self.SearchButton.setStatusTip('Ctrl+F')

		if self.CurrentMonth[1] == '01':
			self.CurrentMonth[1] = '12'
			self.CurrentMonth[0] = str(int(self.CurrentMonth[0]) - 1)

		else:
			self.CurrentMonth[1] = str(int(self.CurrentMonth[1]) - 1)

		self.StartCalendar.setCurrentPage(int(self.CurrentMonth[0]), int(self.CurrentMonth[1]))
		self.EndCalendar.setCurrentPage(int(self.CurrentMonth[0]), int(self.CurrentMonth[1]))
		self.StartCalendar.setSelectedDate(QtCore.QDate.fromString('-'.join(self.CurrentMonth + ['01']), 'yyyy-MM-dd'))
		self.EndCalendar.setSelectedDate(QtCore.QDate.fromString('-'.join(self.CurrentMonth + [str(self.LastDay)]), 'yyyy-MM-dd'))
		
		self.WriteButton.clicked.connect(self.Write)
		self.SearchButton.clicked.connect(self.Search)
		self.SelectAll.stateChanged.connect(self.Select)
		self.SingleTimer.timeout.connect(self.SingleShot)
		self.WriteTimer.timeout.connect(self.Write)
		self.AnimationTimer.timeout.connect(self.Animation)

		if path.exists(secretpath):
			self.InformationLabel.setText('Retrieving profiles from Google Analytics')
			self.InformationLabel.setStyleSheet('color: green')

			self.SingleTimer.start(200)

		else:
			self.AnimationTimer.start(200)
			self.InformationLabel.setText('The client_secrets.json file is missing from the program installation directory')
			self.InformationLabel.setStyleSheet('color: red')


	def CompanyInformation(self):

		global companysite, cellphone
		
		if not path.exists(infopath):
			companysite, ok = QtWidgets.QInputDialog.getText(self, 'Input Dialog', 'Company Website:                                         ')
			
			if ok:
				cellphone, ko = QtWidgets.QInputDialog.getText(self, 'Input Dialog',   'Cellphone:                                             ')

				if ko:
					with open(infopath, 'w') as companyinfo:
						companyinfo.write('%s: | :%s' % (str(companysite), str(cellphone)))

				else:
					raise SystemExit

			else:
				raise SystemExit

		else:
			with open(infopath, 'r') as companyinfo:
				readinfo = companyinfo.read()
				companysite, cellphone = readinfo.split(': | :')


	def SingleShot(self):

		if not self.processgoing:
			sleep(0.25)
			self.CompanyInformation()
			self.FetchProcess.start()
			self.processgoing = True

		try:
			parameters = self.pipe.get(False)

			self.processgoing = False

			self.setDisabled(False)
			self.SingleTimer.stop()
			self.AnimationLabel.hide()
			self.InformationLabel.setText(' ')

			if isinstance(parameters, str):
				self.InformationLabel.setText(parameters)
				self.InformationLabel.setStyleSheet('color: red')

			elif isinstance(parameters, list):
				n = 0

				for each in parameters:
					self.item = QtWidgets.QTableWidgetItem(each[0])
					self.SiteList.insertRow(n)
					self.SiteList.setVerticalHeaderItem(n, self.item)

					for m in range(0, len(each)):
						self.item = QtWidgets.QTableWidgetItem(each[m])
						self.item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
						self.SiteList.setItem(n, m, self.item)

					n += 1
				
		except Empty:
			if not self.FetchProcess.is_alive():
				self.SingleTimer.stop()
				self.AnimationLabel.hide()
				self.InformationLabel.setText('Could not retrieve information from Google Analytics servers')
				self.InformationLabel.setStyleSheet('color: red')
				
		self.Animation()


	def Write(self):

		if not self.processgoing:
			self.InformationLabel.setText(' ')
			self.WriteProgress.setValue(0)

			if not self.SiteList.selectionModel().selectedRows():
				self.InformationLabel.setText('At least one row needs to be selected')
				self.InformationLabel.setStyleSheet('color: red')

			else:
				profilesinfo = []

				for selectedrow in self.SiteList.selectionModel().selectedRows():
					profile_id = self.SiteList.item(selectedrow.row(), 0).text()
					website = self.SiteList.item(selectedrow.row(), 2).text()
					startdate = str(self.StartCalendar.selectedDate()).split('(')[-1].strip(')').split(', ')
					enddate = str(self.EndCalendar.selectedDate()).split('(')[-1].strip(')').split(', ')

					for i in range(0, len(startdate)):
						startdate[i] = startdate[i].zfill(2)

					for i in range(0, len(enddate)):
						enddate[i] = enddate[i].zfill(2)

					startdate = '-'.join(startdate)
					enddate = '-'.join(enddate)

					profilesinfo.append([profile_id, website, startdate, enddate])

				self.FetchProcess = Thread(target = FilesWriter, args = (self.pipe, profilesinfo))
				self.lengthbar = 100 / len(profilesinfo)
				self.processgoing = True

				self.WriteTimer.start(90)
				self.FetchProcess.start()
				self.AnimationLabel.show()
				self.WriteButton.setDisabled(True)

		else:
			try:
				string = self.pipe.get(False)

				if 'Fetching' not in string and self.prevdone:
					self.WriteProgress.setValue(round(self.WriteProgress.value() + self.lengthbar))

				if string == 'done':
					self.WriteTimer.stop()
					self.AnimationLabel.hide()

					self.WriteProgress.setValue(100)
					self.InformationLabel.setText('Completed')
					self.WriteButton.setDisabled(False)

					self.processgoing = 0
					self.prevdone = False

				else:
					self.InformationLabel.setText(string)
					self.InformationLabel.setStyleSheet('color: green')

					if not 'Fetching' in string:
						self.AnimationLabel.hide()
						self.prevdone = True
					
			except Empty:
				if not self.FetchProcess.is_alive():
					self.WriteTimer.stop()
					self.AnimationLabel.hide()
					self.InformationLabel.setText('Could not retrieve information from Google Analytics servers')
					self.InformationLabel.setStyleSheet('color: red')

			self.Animation()


	def Select(self):

		items = self.SiteList.selectedItems()

		if self.SelectAll.checkState():
			for x in range(0, self.SiteList.rowCount()):
				if not self.SiteList.item(x, 0) in items:
					self.SiteList.setCurrentCell(x, 0)

		else:
			for x in range(0, self.SiteList.rowCount()):
				if self.SiteList.item(x, 0) in items:
					self.SiteList.setCurrentCell(x, 0)


	def Search(self):

		self.InformationLabel.setText(' ')

		try: # Procura a primeira linha com determinados caracteres
			if not self.prevtext == self.SearchText.text():
				raise AttributeError

			self.SiteList.setCurrentCell(self.searcheditems[0].row(), 0)
			del self.searcheditems[0]

		except (AttributeError, IndexError): # Se for procurado outra vez, irá procurar a seguir
			self.searcheditems = self.SiteList.findItems(self.SearchText.text(), QtCore.Qt.MatchContains)
			self.prevtext = self.SearchText.text()

			if self.searcheditems:
					self.SiteList.setCurrentCell(self.searcheditems[0].row(), 0)
					del self.searcheditems[0]

			else:
				self.InformationLabel.setText('No item matches the criteria')
				self.InformationLabel.setStyleSheet('color: orange')


	def Animation(self):

		if self.count >= len(self.charlist) - 1:
			self.count = 0

		self.AnimationLabel.setText(self.charlist[self.count])
		self.count += 1


	def closeEvent(self, event):
			
			reply = QtWidgets.QMessageBox.question(self, 'Analytical', "\nExit the program?\n", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)

			if reply == QtWidgets.QMessageBox.Yes:
				event.accept()

			else:
				event.ignore()


class Statistics(object):

	def WorksheetGenerator(wbook, results, title, startdate, enddate, yeardate):

		company = results.get('profileInfo').get('profileName')
		filename = company + ' ' + startdate + ' ' + enddate + '.xlsx'
		filepath = path.join(folderpath, filename)
			
		header = [h['name'][3:] for h in results.get('columnHeaders')]
		header.insert(0, u'')

		x = 0

		for y in range(1, len(header)):
			header[y] = header[y].capitalize()

		if results.get('rows', []):
			if title == 'Sessões (Gerais)':
				percentage = True
				c, sessionsum, usersum, pagevsum, upagevsum, monthlysessiond, monthlytimep, rejectionrate = 0, 0, 0, 0, 0, 0, 0, 0

				wsheet = wbook.active
				wsheet.title = title

				header.append('Pagepersession')
				wsheet.append([u''])
				wsheet.append(header)
				sessionlist = []

				for data in results.get('rows', []):
					data[0] = data[0][:4] + '-' + data[0][4:6] + '-' + data[0][6:8] # Rearrange date
					sessionsum += int(data[1]) # Overall session sum
					usersum += int(data[2]) # Overall user sum
					pagevsum += int(data[3]) # Overall pageviews sum
					upagevsum += int(data[4]) # Overall unique pageviews sum
					monthlysessiond += float(data[5]) # Monthly session duration
					monthlytimep += float(data[6]) # Monthly time on page
					rejectionrate += float(data[7]) # Number of single page sessions (30 minutes without interacting)

					data[5] = TimeFormat(data[5])
					data[6] = TimeFormat(data[6])
					data[7] = float(data[7]) / 100

					sessionlist.append(int(data[1]))

					IntegerFormat(data, percentage)
					data.insert(0, u'')
					wsheet.append(data)

					c += 1

				monthlysessiond = TimeFormat(monthlysessiond / c)
				monthlytimep = TimeFormat(monthlytimep / c)

				try:
					total = [u'', 'Total', sessionsum, usersum, pagevsum, upagevsum]
					avg = [u'', 'Média', round(sessionsum / c), round(usersum / c), round(pagevsum / c), round(upagevsum / c), monthlysessiond, 
						monthlytimep, rejectionrate / c / 100, round(pagevsum / sessionsum, 1)]

				except ZeroDivisionError:
					total = [u'', 'Total', '0', '0', '0', '0']
					avg = [u'', 'Média', '0', '0', '0', '0', '0', '0', '0', '0']

				wsheet.append([u''])
				wsheet.append(total)
				wsheet.append(avg)
				PDFlist.extend([sessionlist, sessionsum, usersum, avg[2], pagevsum, avg[9], monthlysessiond, str(round(float(avg[8]) * 100)) +'%'])
				
				Statistics.ChartConstruction(wsheet, ['col', 16], 'Sessões de ' + months[enddate.split('-')[1]], 'M10', 3, 2, 3, 33, False)
				wsheet.cell('B34').font = headerfont
				wsheet.cell('B35').font = headerfont
				wsheet.cell('B36').font = headerfont
				

			elif title == 'Sessões (Países)' or title == 'Sessões (Cidades)':
				percentage = True
				wsheet = wbook.create_sheet()
				wsheet.title = title
				header[3] = 'Pagepersession'

				wsheet.append([u''])
				wsheet.append(header)

				clist, csessions, cpagepersession, crejectionrate = [], [], [], []

				for data in results.get('rows', []):
					data[2] = str(round(int(data[2]) / int(data[1]), 2))
					data[3] = float(data[3]) / 100

					if not len(clist) >= 10 and data[0] != '(not set)':
						clist.append(data[0])
						csessions.append(int(data[1]))
						cpagepersession.append(float(data[2]))
						crejectionrate.append(str(round(float(data[3]) * 100)) + '%')

					IntegerFormat(data, percentage)
					data.insert(0, u'')
					wsheet.append(data)

				Statistics.ChartConstruction(wsheet, ['pie', None], '10 ' + title.split(' ')[1].strip(')(') + ' com Mais Sessões', 'H4', 3, 2, 3, 12, [2, 3, 12])
				PDFlist.extend([clist, csessions, cpagepersession, crejectionrate])


			elif title == 'Sessões (Diárias)':
				percentage = False
				wsheet = wbook.create_sheet()
				wsheet.title = title

				header = [u'', 'Hours', 'Sessions']
				hourlydict = {}

				wsheet.append([u''])
				wsheet.append(header)
					
				for data in results.get('rows', []):
					if int(data[0][-2:]) in hourlydict:
						hourlydict[int(data[0][-2:])] += int(data[1])

					else:
						hourlydict[int(data[0][-2:])] = int(data[1])

				hourlist, hsessions = [], []

				for hour, users in hourlydict.items():
					data = [u'', hour, users]
					hourlist.append(str(hour))
					hsessions.append(users)
					IntegerFormat(data, percentage)
					wsheet.append(data)

				Statistics.ChartConstruction(wsheet, ['lin', 13], 'Sessões Diárias', 'F7', 3, 2, 3, 25, False)
				PDFlist.extend([hourlist, hsessions])


			elif title == 'Fontes de Acesso' or title == 'Palavras-Chaves':
				percentage = True
				wsheet = wbook.create_sheet()
				wsheet.title = title

				header.append('Sessionsrate')
				wsheet.append([u''])
				wsheet.append(header)

				for sessionsum in results.get('rows', []):
					x += int(sessionsum[1])

				flist, fsessions, fpercentage = [], [], []

				for data in results.get('rows', []):
					data.append(float(data[1]) / x)

					if not len(flist) >= 10:
						flist.append(data[0])
						fsessions.append(int(data[1]))

					IntegerFormat(data, percentage)
					data.insert(0, u'')
					wsheet.append(data)

				for each in fsessions:
					fpercentage.extend([str(round(each / sum(fsessions) * 100)) + '%'])

				PDFlist.extend([flist, fsessions, fpercentage])


			elif title == 'Páginas':
				percentage = True
				wsheet = wbook.create_sheet()
				wsheet.title = title

				header.append('Sessionsrate')
				wsheet.append([u''])
				wsheet.append(header)

				for sessionsum in results.get('rows', []):
					x += int(sessionsum[1])

				plist, ppageview, ppercentage, pavgtimeonpage = [], [], [], []

				for data in results.get('rows', []):
					data[2] = TimeFormat(data[2])

					plist.append(data[0])
					ppageview.append(int(data[1]))
					pavgtimeonpage.append(data[2])

					IntegerFormat(data, percentage)
					data.insert(0, u'')
					data.insert(4, float(data[2]) / x)
					wsheet.append(data)

				for each in ppageview:
					ppercentage.append(str(round(each / sum(ppageview) * 100)) + '%')

				PDFlist.extend([plist, ppageview, ppercentage, pavgtimeonpage])

			elif title == 'Tracking de Páginas':

				pagetrack = []

				for data in results.get('rows', [])[:6]:
					if data[0] != data[1] and data[1] != '(not set)':
						pagetrack.append(data[:2])

				PDFlist.append(pagetrack)

				return


			elif title == 'Sessões (Anuais)':
				percentage = False
				wsheet = wbook.create_sheet()
				wsheet.title = title

				wsheet.append([u''])
				wsheet.append(header)

				for sessionsum in results.get('rows', []):
					x += int(sessionsum[1])

				month = yeardate.split('-')[1]
				ordlist = results.get('rows', [])[int(month) - 1:] + results.get('rows', [])[:int(month) - 1] # Improves Google monthly display
				
				ysessions, monthlist = [], []

				for data in ordlist:
					IntegerFormat(data, percentage)
					ysessions.append(data[1])

					if month == '13':
						month = '01'

					wsheet.append([u'', months[month], data[1]])
					monthlist.append(months[month][:3] + '.')
					month = str(int(month) + 1).zfill(2)

				data = [u'', 'Média', round(int(x) / 12)]
				wsheet.append([u''])
				wsheet.append(data)
				wsheet.cell('B16').font = headerfont

				Statistics.ChartConstruction(wsheet, ['lin', 13], 'Sessões Anuais', 'F3', 3, 2, 3, 14, [2, 3, 14])
				PDFlist.extend([ysessions, round(int(x) / 12), monthlist])
					

			if percentage == True:
				for i in range(3, len(results.get('rows', [])) + 6):
					wsheet[chr(len(data) + 64) + str(i)].number_format = '0%'

			Statistics.HeaderFormat(wsheet, len(header))
			Statistics.WorksheetFormat(wsheet, len(data), len(results.get('rows', [])))

			if title == 'Sessões (Anuais)':
				wbook.save(filename = filepath)

				del wbook

				Report.PageGenerator(filename.rpartition('.')[0], enddate)


	def HeaderFormat(wsheet, header):

		x = 66

		while not x - 65 == header:
			wsheet.cell(str(chr(x)) + '2').font = headerfont
			wsheet.cell(str(chr(x)) + '2').alignment = align
			wsheet.cell(str(chr(x)) + '2').border = border
			x += 1

		return wsheet


	def WorksheetFormat(wsheet, columns, lines):

		x, y = 66, 3

		while not y - 6 == lines:
			while not x - 66 == columns:
				wsheet.cell(str(chr(x)) + str(y)).alignment = align
				x += 1

			x = 66
			y += 1

		return wsheet


	def ChartConstruction(wsheet, typechart, name, location, minc, minr, maxc, maxr, categories):

		if typechart[0] == 'lin':
			chart = LineChart()

		elif typechart[0] == 'pie':
			chart = PieChart()
			chart.splitType = 'val'

		elif typechart[0] == 'col':
			chart = BarChart()
			chart.type = typechart[0]
			chart.shape = 4

		data = Reference(wsheet, min_col = minc, min_row = minr, max_col = maxc, max_row = maxr)
		chart.title = name
		chart.style = typechart[1]
		chart.add_data(data, titles_from_data = True)

		if categories:
			chart.set_categories(Reference(wsheet, min_col = categories[0], min_row = categories[1], max_row = categories[2]))

		wsheet.add_chart(chart, location)
		
		return wsheet


class Report(object):

	def PageGenerator(filename, enddate):

		global doc, story

		info, sessionsperday, sessionsum, usersum, avgsessions, pagevsum, avgpageviews, monthlysessiond, rejectionrate, accesslist, acesssessions, accesspercentage, searchlist, searchsessions, searchpercentage, countrylist, countrysessions, countrypages, countryrejections, citylist, citysessions, citypages, cityjections, plist, ppageview, ppercentage, pavgtimeonpage, hourslist, hoursessions, pagetrack, ysessions, avgvisitspermonth, monthlist = PDFlist

		filepath = path.join(folderpath, filename + '.pdf')

		doc = SimpleDocTemplate(filepath)
		style = getSampleStyleSheet()['Normal']
		story = [Spacer(1, 2 * inch)]
		monthdate = [str(x) for x in range(1, int(enddate.split('-')[2]) + 1)]
		maxhour = hourslist[hoursessions.index(max(hoursessions))]
		temphours = hourslist
		tempsessions = hoursessions

		del temphours[tempsessions.index(max(tempsessions))]
		del tempsessions[tempsessions.index(max(tempsessions))]

		secmaxhour = temphours[tempsessions.index(max(tempsessions))]

		if len(accesslist) > 1:
			searchindex = accesslist[0] if '(direct)' not in accesslist[0] else accesslist[1]

		else:
			searchindex = accesslist[0]

		story.append(PageBreak())
		
		story.append(Paragraph('<br/><br/><br/><br/><br/><br/>', style))
		story.append(Paragraph('<br/><br/><br/><br/><br/><br/>', style))
		story.append(Paragraph('<para alignment="justify"><a href = page3.html#0>1. Estatísticas Globais . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . &nbsp;3</a></para>', pH))
		story.append(Paragraph('<para alignment="justify"><a href = page3.html#1>1.1. Sessões . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . 3</a></para>', sH))
		story.append(Paragraph('<para alignment="justify"><a href = page3.html#2>1.2. Pageviews . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . &nbsp;3</a></para>', sH))
		story.append(Paragraph('<para alignment="justify"><a href = page3.html#3>1.3. Tempo Médio no Website . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . &nbsp;3</a></para>', sH))
		story.append(Paragraph('<para alignment="justify"><a href = page3.html#4>1.4. Taxa de Rejeição . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . &nbsp;3</a></para>', sH))
		story.append(Paragraph('<para alignment="justify"><a href = page4.html#2>2. Geografia de Visitantes . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . 4</a></para>', pH))
		story.append(Paragraph('<para alignment="justify"><a href = page5.html#3>3. Sessões por Hora do Dia . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . &nbsp;5</a></para>', pH))
		story.append(Paragraph('<para alignment="justify"><a href = page5.html#4>4. Fontes de Acesso ao Website . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . &nbsp;5</a></para>', pH))
		story.append(Paragraph('<para alignment="justify"><a href = page6.html#5>5. Palavras-Chave nas Pesquisas nos Motores de Busca . . . . . . . . . . . . . . . . . . . . . 6</a></para>', pH))
		story.append(Paragraph('<para alignment="justify"><a href = page6.html#6>6. Histórico da Visita - Reconstrução das Sessões Tipo . . . . . . . . . . . . . . . . . . . . . . 6</a></para>', pH))
		story.append(Paragraph('<para alignment="justify"><a href = page7.html#7>7. Áreas Visitadas do Website - Ranking por Pageview . . . . . . . . . . . . . . . . . . . . . &nbsp;7</a></para>', pH))
		story.append(Paragraph('<para alignment="justify"><a href = page7.html#8>8. Número de Sessões ao Longo do Ano . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . &nbsp;7</a></para>', pH))
		story.append(Paragraph('<para alignment="justify"><a href = page8.html#9>9. Conclusão . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . 8</a></para>', pH))
		story.append(PageBreak())

		story.append(Paragraph('<br/><br/><br/><br/><br/>', style))
		story.append(Paragraph('<a name = page3.html#0></a>1. Estatísticas Globais', pT))
		story.append(Paragraph('<a name = page3.html#1></a>1.1. Sessões', sT))
		
		draw = Drawing((defaultPageSize[0] - 400) / 2, 200)
		chart = VerticalBarChart()
		chart.width = 400
		chart.height = 170
		chart.data = [sessionsperday]
		chart.categoryAxis.categoryNames = [str(x) for x in range(1, int(enddate.split('-')[2]) + 1)]
		chart.bars[0].fillColor = colors.HexColor('#47d147')
		chart.bars[0].strokeColor = colors.white
		draw.add(chart)
		story.append(draw)

		story.append(Paragraph('<br/><br/>', style))
		story.append(Paragraph('<para alignment="center">No presente mês, o website obteve <b>%s</b> visitas, das quais se identificaram <b>%s</b> visitantes únicos. <br/>O website teve uma média de <b>%s</b> visitas por dia.</para>' % (sessionsum, usersum, avgsessions), style))
		story.append(Paragraph('<br/><br/>', style))
		story.append(Paragraph('<a name = page3.html#2></a>1.2. Pageviews', sT))
		story.append(Paragraph('<br/>', pT))
		story.append(Paragraph('Das <b>%s</b> visitas verificadas, foram visualizadas <b>%s</b> páginas.<br/>Em cada visita foram consultadas, em média, <b>%s</b> páginas.' % (sessionsum, pagevsum, avgpageviews), style))
		story.append(Paragraph('<br/><br/>', style))
		story.append(Paragraph('<a name = page3.html#3></a>1.3. Tempo Médio no Website', sT))
		story.append(Paragraph('<br/>', pT))
		story.append(Paragraph('O tempo médio de uma visita no corrente mês foi de <b>%s</b>.' % monthlysessiond, style))
		story.append(Paragraph('<br/><br/>', style))
		story.append(Paragraph('<a name = page3.html#4></a>1.4. Taxa de Rejeição', sT))
		story.append(Paragraph('<br/>', pT))
		story.append(Paragraph('No mês de Outubro houve uma taxa de rejeição<b>&sup1;</b> de <b>%s</b>.' % rejectionrate, style))
		story.append(Paragraph('<br/><br/><br/><br/>', style))
		story.append(Paragraph('&sup1; - Percentagem de utilizadores que consultaram apenas uma página', sS))
		story.append(PageBreak())

		story.append(Paragraph('<br/><br/><br/><br/><br/>', style))
		story.append(Paragraph('<a name = page4.html#2></a>2. Geografia de Visitantes', pT))
		story.append(Paragraph('<br/><br/><br/>', style))

		table = Report.TableFormat(['Países', 'Sessões', 'Páginas/Sessões', 'Taxa de Rejeição'], countrylist, countrysessions, countrypages, countryrejections)
		
		story.append(table)
		story.append(Paragraph('<br/><br/><br/><br/><br/><br/>', style))

		table = Report.TableFormat(['Cidades', 'Sessões', 'Páginas/Sessões', 'Taxa de Rejeição'], citylist, citysessions, citypages, cityjections)
		
		story.append(table)
		story.append(PageBreak())

		story.append(Paragraph('<br/><br/><br/><br/><br/>', style))
		story.append(Paragraph('<a name = page5.html#3></a>3. Visitas por Hora do Dia', pT))
		story.append(Paragraph('<br/><br/>', style))

		draw = Drawing((defaultPageSize[0] - 400) / 2, 200)
		chart = HorizontalLineChart()
		chart.width = 400
		chart.height = 170
		chart.data = [hoursessions]
		chart.categoryAxis.categoryNames = hourslist #[0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23]
		draw.add(chart)
		story.append(draw)

		story.append(Paragraph('<br/><br/>', style))
		story.append(Paragraph('<para align="center">O gráfico acima apresenta o número total de visitas distribuídas ao longo das 24 horas.</para>', style))
		story.append(Paragraph('<br/><br/>', style))
		story.append(Paragraph('<a name = page5.html#4></a>4. Fontes de Acesso ao Website', pT))
		story.append(Paragraph('<br/><br/><br/>', style))

		table = Report.TableFormat(['Origem', 'Visitas', '% Visitas'], accesslist, acesssessions, accesspercentage)

		story.append(table)
		story.append(PageBreak())

		story.append(Paragraph('<br/><br/><br/><br/><br/>', style))
		story.append(Paragraph('<a name = page6.html#5></a>5. Palavras-Chave nas Pesquisas nos Motores de Busca', pT))
		story.append(Paragraph('<br/><br/><br/>', style))

		table = Report.TableFormat(['Palavras-Chave&sup2;', 'Visitas', '% Visitas'], searchlist, searchsessions, searchpercentage)

		story.append(table)
		story.append(Paragraph('<br/><br/><br/>', style))
		story.append(Paragraph('<a name = page6.html#6></a>6. Histórico da Visita - Reconstrução das Sessões Tipo', pT))
		story.append(Paragraph('<br/><br/><br/>', style))
		story.append(Paragraph('<u>Caminho mais útil para os utilizadores</u>:', style))
		story.append(Paragraph('<br/><br/>', style))

		if pagetrack:
			story.append(Paragraph('<b>1)</b> %s <b>></b> %s' % (pagetrack[0][0], pagetrack[0][1]), style))

		else:
			story.append(Paragraph('<i>Sem caminhos úteis disponíveis</i>', style))

		story.append(Paragraph('<br/><br/><br/>', style))
		story.append(Paragraph('<u>Outros caminhos úteis</u>:', style))
		story.append(Paragraph('<br/><br/>', style))

		if len(pagetrack) > 2:
			c = 2

			for route in pagetrack[1:]:
				story.append(Paragraph('<b>%s)</b> %s <b>></b> %s' % (c, route[0], route[1]), style))
				c += 1

		else:
			story.append(Paragraph('<i>Sem caminhos úteis disponíveis</i>', style))

		story.append(Paragraph('<br/><br/>', style))
		story.append(Paragraph('&sup2; - As palavras-chave mais utilizadas para encontrar o website através de motores de busca como o Google, Bing e outros', sS))
		story.append(PageBreak())

		story.append(Paragraph('<br/><br/><br/><br/><br/>', style))
		story.append(Paragraph('<a name = page7.html#7></a>7. Áreas Visitadas do Website - Ranking por Pageview', pT))
		story.append(Paragraph('<br/><br/><br/>', style))

		table = Report.TableFormat(['Título da Página', 'Visualizações', '% Visualizações', 'Média (t)'], plist, ppageview, ppercentage, pavgtimeonpage)
		
		story.append(table)
		story.append(Paragraph('<br/><br/><br/>', style))
		story.append(Paragraph('<a name = page7.html#8></a>8. Número de Sessões ao Longo do Ano', pT))
		story.append(Paragraph('<br/><br/>', style))

		draw = Drawing((defaultPageSize[0] - 400) / 2, 200)
		chart = VerticalBarChart()
		chart.width = 400
		chart.height = 170
		chart.data = [ysessions]
		chart.categoryAxis.categoryNames = list(monthlist)
		chart.bars[0].fillColor = colors.HexColor('#80ccff', hasAlpha = True)
		chart.bars[0].strokeColor = colors.white
		draw.add(chart)
		story.append(draw)

		story.append(Paragraph('<br/><br/>', style))
		story.append(Paragraph('<para alignment="center">Verifica-se que o website obteve em média <b>%s</b> visitas por mês.</para>' % avgvisitspermonth, style))
		story.append(PageBreak())

		story.append(Paragraph('<br/><br/><br/><br/><br/>', style))
		story.append(Paragraph('<a name = page8.html#9></a>9. Conclusão', pT))
		story.append(Paragraph('<br/><br/><br/>', style))

		primeday = monthdate[sessionsperday.index(max(sessionsperday))]
		toyear, tomonth = enddate.split('-')[:2]
		toweekday = weekdays[weekday(int(toyear), int(tomonth), int(primeday))]

		story.append(Paragraph('Em resumo, poderá afirmar-se que o dia mais visitado foi <b>%s de %s (%s)</b> com <b>%s</b> visitas, sendo que o maior número de visitas teve origem de <b>%s</b>.' % (primeday, PDFlist[0][1], toweekday, max(sessionsperday), citylist[citysessions.index(max(citysessions))]), style))
		story.append(Paragraph('<br/><br/>', style))
		story.append(Paragraph('Por outro lado, <b>%s</b> dos acessos ao website surge através do <b>%s</b> e a palavra-chave mais utilizada foi <b>%s</b>.' % (accesspercentage[accesslist.index(searchindex)], searchindex, searchlist[searchsessions.index(max(searchsessions))]), style)) 
		story.append(Paragraph('<br/><br/>', style))

		if len(plist) > 1:
			primepage = plist[0] if '(not set)' not in plist[0] else plist[1]

		else:
			primepage = plist[0]

		story.append(Paragraph('A página mais acedida, além da página inicial, foi a <b>%s</b>, verificando-se que os períodos diários com maior tráfego no website foram às <b>%sh</b> e às <b>%sh</b>.' % (primepage, maxhour, secmaxhour), style))

		doc.build(story, onFirstPage = Report.FirstPage, onLaterPages = Report.LaterPages)


	def FirstPage(canvas, doc):

		height = 120

		canvas.saveState()

		heightcenter = ImageFormat(images['logo'])
		widthcenter = images['logo']['width']
		canvas.drawImage(images['logo']['path'], (defaultPageSize[0] - widthcenter) / 2, defaultPageSize[1] - heightcenter - height, widthcenter, preserveAspectRatio = True, mask = 'auto')

		height += heightcenter

		heightcenter = ImageFormat(images['background'])
		widthcenter = images['background']['width']
		canvas.drawImage(images['background']['path'], (defaultPageSize[0] - widthcenter) / 2, (defaultPageSize[1] - heightcenter - (height / 2) - (0.75 * inch)) / 2, widthcenter, preserveAspectRatio = True, mask = 'auto')

		canvas.setFont('Calibri', 32)
		canvas.drawCentredString(defaultPageSize[0] / 2, defaultPageSize[1] / 2 + 15, "Relatório Mensal de")
		canvas.drawCentredString(defaultPageSize[0] / 2, defaultPageSize[1] / 2 - 20, '%s %s' % (PDFlist[0][1], PDFlist[0][2]))
		canvas.setFillColorRGB(0, 0, 0.6)

		canvas.drawCentredString(defaultPageSize[0] / 2, defaultPageSize[1] / 2 - 55, PDFlist[0][0])

		canvas.setFont('Times-Roman', 9)
		canvas.setFillColorRGB(0, 0, 0) # (0, 0.7, 0.9)
		canvas.drawCentredString(defaultPageSize[0] / 2, 0.75 * inch, "%s | Telf: %s" % (companysite, cellphone))
		canvas.restoreState()


	def LaterPages(canvas, doc):

		canvas.saveState()

		heightcenter = ImageFormat(images['header'])
		widthcenter = images['header']['width']
		canvas.drawImage(images['header']['path'], (defaultPageSize[0] - widthcenter) / 2, defaultPageSize[1] - 170, widthcenter, preserveAspectRatio = True, mask = 'auto')

		heightcenter = ImageFormat(images['footer'])
		widthcenter = images['footer']['width']
		canvas.drawImage(images['footer']['path'], (defaultPageSize[0] - widthcenter) / 2, 40, widthcenter, preserveAspectRatio = True, mask = 'auto')

		canvas.setFont('Times-Roman', 9)
		canvas.drawString(70, 1.15 * inch, "Relatório Mensal de Visitas - {}".format(PDFlist[0][1]))
		canvas.drawRightString(defaultPageSize[0] - 70, 1.15 * inch, "{} {} | Página {}".format(PDFlist[0][2], date.today().strftime('%d-%m-%Y'), doc.page))
		canvas.drawRightString(defaultPageSize[0] - 70, 1.15 * inch - 15, "{} | Telf: {}".format(companysite, cellphone))
		canvas.restoreState()


	def TableFormat(title, *args):

		y, data = 0, [[]]

		style = getSampleStyleSheet()['Normal']

		for each in title:
			data[0].extend([Paragraph('<para alignment="center"><b>%s</b></para>' % each, style)])

		for xgroup in args[0]:
			ygroup = []
			x = 0

			while not x >= len(args):
				if len(str(args[x][y])) > 40:
					args[x][y] = args[x][y][:41] + '(...)'

				ygroup.append(args[x][y])

				x += 1

			data.append(ygroup)
				
			y += 1

		if len(data) < 10:
			while not len(data) == 11:
				data.append([' '])

		elif len(data) > 10:
			data = data[:11]

		table = Table(data, style = [('ALIGN', (1, 1), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
			('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#29a3a3')), 
			('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#c2f0f0')),
			('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#c2f0f0')),
			('BACKGROUND', (0, 6), (-1, 6), colors.HexColor('#c2f0f0')),
			('BACKGROUND', (0, 8), (-1, 8), colors.HexColor('#c2f0f0')),
			('BACKGROUND', (0, 10), (-1, 10), colors.HexColor('#c2f0f0')),
		])

		return table


def OnLoadWorker(pipe):

	try:
		service, flags = sample_tools.init(argv, 'analytics', 'v3', __doc__, curdir, scope = 'https://www.googleapis.com/auth/analytics.readonly')
		
		profiles = GetProfileInfo(service)

		parameters = []
				
		for profileID in profiles:
			newsite = profiles[profileID][1]

			if 'http' not in newsite[:4]:
				newsite = 'http://' + profiles[profileID][1]

			if '/' not in newsite[-1]:
				newsite = newsite + '/'

			parameters.append([profileID, profiles[profileID][0], newsite])

		pipe.put(parameters, block = True)

	except TypeError as error: # Handle errors in constructing a query
		pipe.put('There was an error in constructing your query', block = True)

	except HttpError as error: # Handle API errors
		pipe.put('There was an API error : %s - %s' % (error.resp.status, error._get_reason()), block = True)

	except AccessTokenRefreshError: # Handle Auth errors
		pipe.put('The credentials have been revoked or expired, please re-run the application to re-authorize', block = True)

	except FileNotFoundError as e: # Handles Name errors
		pipe.put('A website name has unexpected punctuation or path to files has changed. %s' % e, block = True)

	except PermissionError:
		pipe.put('The file in question needs to be closed in order to be updated', block = True)

	except Exception as e:
		pipe.put('%s' % e, block = True)

	except KeyboardInterrupt:
		raise SystemExit
		

def GetProfileInfo(service):

	profilesinfo = OrderedDict()

	accounts = service.management().accounts().list().execute()

	if accounts.get('items'):
		for accountNumber in range(0, len(accounts.get('items'))):
			firstAccountId = accounts.get('items')[accountNumber].get('id')
			webproperties = service.management().webproperties().list(accountId = firstAccountId).execute()

			if webproperties.get('items'):
				for webId in webproperties.get('items'):
					profiles = service.management().profiles().list(accountId = firstAccountId, webPropertyId = webId.get('id')).execute()

					if profiles.get('items'):
						profilesinfo[profiles.get('items')[0].get('id')] = (profiles.get('items')[0].get('name'), profiles.get('items')[0].get('websiteUrl'))
						sleep(0.1)

		if webproperties.get('items'):
			return profilesinfo

	return None


def FilesWriter(pipe, info):

	global PDFlist

	pipe.put('Fetching information from Google Analytics', block = True)

	service, flags = sample_tools.init(argv, 'analytics', 'v3', __doc__, curdir, scope = 'https://www.googleapis.com/auth/analytics.readonly')

	for selectedrow in info:
		PDFlist = [[selectedrow[1], months[selectedrow[3].split('-')[1]], selectedrow[3].split('-')[0]]]

		wbook = workbook.Workbook()

		counter = 0

		for stats in group:
			pstartdate = selectedrow[2]

			if stats == 'yearly':
				splitdate = selectedrow[3].split('-')
				pstartdate = str(int(splitdate[0]) - 1).zfill(2) + '-01-01'

				if not splitdate[1] == '12':
					pstartdate = str(int(splitdate[0]) - 1).zfill(2) + '-' + str(int(splitdate[1]) + 1).zfill(2) + '-01'

			if not len(group[stats]) == 3:
				group[stats].append(None)

			results = service.data().ga().get(ids = 'ga:' + selectedrow[0], start_date = pstartdate, end_date = selectedrow[3], dimensions = group[stats][0], 
				metrics = group[stats][1], sort = group[stats][2]).execute()

			if stats == 'session':
				company = results.get('profileInfo').get('profileName')
				pipe.put('Writing %s report and statistics' % company, block = True)

			try:
				Statistics.WorksheetGenerator(wbook, results, titles[counter], selectedrow[2], selectedrow[3], pstartdate)

			except ValueError:
				counter += 1
				break

			counter += 1

		counter = 0

	pipe.put('done', block = True)


def ImageFormat(image):

	x, y = Image.open(image['path']).size
	height = y * image['width'] / x

	return height


def IntegerFormat(data, percentage):

	listlen = len(data)

	if percentage == True:
		listlen = len(data) - 1

	for num in range(1, listlen):
		try:
			data[num] = int(data[num])

		except ValueError:
			pass

	return data


def TimeFormat(seconds): # Converting seconds to time format

	hours, rest = divmod(round(float(seconds)), 3600) 
	minutes, seconds = divmod(rest, 60)

	time = []

	for num in hours, minutes, seconds:
		if len(str(num)) == 1:
			num = '0' + str(num)

		time.append(num)

	return '%s:%s:%s' % (time[0], time[1], time[2])


if __name__ == '__main__':
	if not path.exists(folderpath) and (shell.IsUserAnAdmin() or not frozen):
		makedirs(folderpath)

	app = QtWidgets.QApplication(argv)
	dialog = Interface()
	dialog.show()
	exit(app.exec_())
